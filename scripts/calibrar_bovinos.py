"""Calibra X-13 para producción bovina contra la columna `desest` de Bovinos.xlsx.

Barre mode x td x seasonalma, corre x13as sobre la serie observada (col `valor`) y compara el
d11 resultante contra la referencia (col `desest`).

Como aves, la referencia NO se reproduce exacto: la más cercana es add + td1coef + s3x5, con
~0.29% de error medio (3% máx). La referencia se hizo con un método algo distinto del x13as
del repo.

Uso: python scripts/calibrar_bovinos.py
"""
import os
import subprocess
import sys
import tempfile
import datetime as dt

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from etl.core import seasonal as S  # noqa: E402

XLSX = os.path.join(os.path.dirname(__file__), os.pardir,
                    "etl", "datasets", "bovinos", "data", "Bovinos.xlsx")


def load():
    ws = openpyxl.load_workbook(XLSX, data_only=True)["Hoja1"]
    dates, obs, ref = [], [], {}
    for r in range(2, ws.max_row + 1):
        d, v, de = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if not isinstance(d, dt.datetime) or v is None:
            continue
        dd = d.date()
        dates.append(dd)
        obs.append(float(v))
        if de is not None:
            ref[(dd.year, dd.month)] = float(de)
    return dates, obs, ref


def run(x13, dates, obs, ref, mode, td, sma):
    wd = tempfile.mkdtemp(prefix="cal_bov_")
    S._write_spc(os.path.join(wd, "serie.spc"), dates, obs, mode=mode, td=td, seasonalma=sma)
    subprocess.run([x13, "serie"], cwd=wd, capture_output=True, text=True, timeout=120)
    p = os.path.join(wd, "serie.d11")
    if not os.path.isfile(p):
        return None
    d11 = {(d.year, d.month): v for d, v in S._parse_d11(p)}
    common = [k for k in ref if k in d11]
    if not common:
        return None
    pct = [abs(d11[k] - ref[k]) / abs(ref[k]) * 100 for k in common]
    return dict(mode=mode, td=td, sma=sma, arima=S._arima_model(wd, "serie"),
                meanpct=sum(pct) / len(pct), maxpct=max(pct))


def main():
    x13 = S._x13_binary()
    if not x13:
        sys.exit("x13as no encontrado (setear X13PATH)")
    dates, obs, ref = load()
    print(f"serie: {len(obs)} meses {dates[0]:%Y-%m}..{dates[-1]:%Y-%m} | ref: {len(ref)}")
    results = []
    for mode in ("add", "mult", "auto"):
        for td in ("none", "td1coef", "td"):
            for sma in ("s3x5", "s3x3"):
                try:
                    r = run(x13, dates, obs, ref, mode, td, sma)
                except Exception:
                    r = None
                if r:
                    results.append(r)
    results.sort(key=lambda r: r["meanpct"])
    print(f"\n{'mode':5} {'td':8} {'sma':5} {'arima':16} {'meanpct%':>9} {'maxpct%':>9}")
    for r in results:
        print(f"{r['mode']:5} {r['td']:8} {r['sma']:5} {str(r['arima']):16} "
              f"{r['meanpct']:9.4f} {r['maxpct']:9.4f}")


if __name__ == "__main__":
    main()
