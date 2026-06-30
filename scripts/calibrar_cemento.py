"""Verifica/calibra la desest de CEMENTO contra la referencia de Juan.

La referencia está en `Automotriz.xlsx` hoja **Hoja9** (`cemento,xlsb,d11`, 1994-01..2025-10).
Corre x13as sobre el despacho nacional ORIGINAL (de cemento.xlsx), TRUNCADO al mismo span que
Juan (para igualar la "vintage"), y compara el d11 contra Hoja9. Prueba nuestro método (el que
matcheó produccion) y un par de variantes, por si cemento usó settings distintos.

Reusa scripts/calibrar.py (build_spc/run_variant/find_binary/interp_nonpositive).

Uso (server, venv, X13PATH):  python scripts/calibrar_cemento.py
"""
from __future__ import annotations

import os
import statistics as st
import sys
from datetime import date

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
import calibrar as C  # noqa: E402

HERE = os.path.dirname(__file__)
AUTO_XLSX = os.path.abspath(os.path.join(HERE, "..", "etl", "datasets", "automotriz", "data", "Automotriz.xlsx"))
CEM_XLSX = os.path.abspath(os.path.join(HERE, "..", "etl", "datasets", "cemento", "data", "cemento.xlsx"))
OUT_BASE = "/tmp/calibrar_cemento"

# Variantes a comparar contra Hoja9. La #0 es el método que dejamos fijo en el pipeline.
GRID = [
    dict(label="MULT  td1coef + s3x5 + automdl + outlier  [metodo del pipeline]", transform="log", model="automdl", outlier=True, reg="td1coef", seasonalma="s3x5", positive=True),
    dict(label="ADITIVO  td1coef + s3x5 + automdl + outlier", transform="none", model="automdl", outlier=True, mode="add", reg="td1coef", seasonalma="s3x5"),
    dict(label="MULT  sin td1coef (s3x5 + automdl + outlier)", transform="log", model="automdl", outlier=True, seasonalma="s3x5", positive=True),
    dict(label="MULT  sin s3x5 (td1coef + automdl + outlier, filtro msr)", transform="log", model="automdl", outlier=True, reg="td1coef", positive=True),
    dict(label="MULT  td(6 coef) + s3x5 + automdl + outlier", transform="log", model="automdl", outlier=True, reg="td", seasonalma="s3x5", positive=True),
    dict(label="MULT  td1coef + s3x5 + automdl, SIN outlier", transform="log", model="automdl", reg="td1coef", seasonalma="s3x5", positive=True),
]


def read_hoja9(path):
    ws = openpyxl.load_workbook(path, data_only=True)["Hoja9"]
    ref = {}
    for r in ws.iter_rows(values_only=True):
        y, m, v = r[0], r[1], r[2]
        if isinstance(y, int) and isinstance(m, int) and v is not None:
            ref[(y, m)] = float(v)
    return ref


def read_cemento(path):
    ws = openpyxl.load_workbook(path, data_only=True).active
    out = []
    for f, v in ws.iter_rows(values_only=True):
        if f is None or v is None:
            continue
        d = f.date() if hasattr(f, "date") else f
        out.append((date(d.year, d.month, 1), float(v)))
    out.sort()
    return out


def err_vs_ref(series, ref):
    errs = []
    for (y, m), j in ref.items():
        got = series.get((y, m))
        if got is not None:
            errs.append(abs(got - j))
    if not errs:
        return None
    return st.mean(errs), max(errs), len(errs)


def main():
    binary = C.find_binary()
    ref = read_hoja9(AUTO_XLSX)
    ref_max = max(ref)  # (year, month) del último mes de Juan
    cem = [(d, v) for d, v in read_cemento(CEM_XLSX) if (d.year, d.month) <= ref_max]
    dates = [d for d, v in cem]
    orig = [v for d, v in cem]
    print(f"binario: {binary}")
    print(f"ref Juan (Hoja9): {len(ref)} meses, hasta {ref_max[0]}-{ref_max[1]:02d}")
    print(f"original cemento (truncado al span de Juan): {len(orig)} meses {dates[0]}..{dates[-1]}")
    print(f"escala: ref[1994-01]={ref[(1994,1)]:.1f}  original[1994-01]={orig[0]:.1f}\n")

    results = []
    for i, opt in enumerate(GRID):
        values = C.interp_nonpositive(orig) if opt.get("positive") else orig
        workdir = os.path.join(OUT_BASE, str(i))
        try:
            series = C.run_variant(binary, dates, values, opt, workdir)
        except Exception as e:
            print(f"[{i}] {opt['label']:55} ERROR {e}")
            continue
        if not series:
            print(f"[{i}] {opt['label']:55} sin salida (ver {workdir}/serie_err.html)")
            continue
        c = err_vs_ref(series, ref)
        arima = C.arima_from_html(workdir)
        results.append((c[0], c[1], c[2], opt["label"], arima, i))

    print("\n==== CEMENTO vs Hoja9 (orden por err_medio) ====")
    print(f"{'err_medio':>10} {'err_max':>10} {'n':>5}  modelo          variante")
    for me, mx, n, label, arima, i in sorted(results, key=lambda r: r[0]):
        print(f"{me:10.2f} {mx:10.1f} {n:5}  {str(arima):14}  [{i}] {label}")
    if results:
        best = min(results, key=lambda r: r[0])
        print(f"\nMEJOR: [{best[5]}] {best[3]}  (err_medio={best[0]:.2f})")
        print("OJO: el original de Juan dice 'xlsb' -> si difiere de cemento.xlsx, parte del "
              "error es por datos de entrada distintos, no por el método.")


if __name__ == "__main__":
    main()
