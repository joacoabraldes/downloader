"""Calibra X-13 para acero crudo contra la columna `desest` de Acero.xlsx.

Barre mode x td x seasonalma, corre x13as sobre la serie observada (col `valor`) y compara
el d11 resultante contra la referencia (col `desest`). Sirve para (re)confirmar los
parámetros del cuadro (etl/series_desest.toml, bloque [acero]).

La variante ganadora reproduce la referencia con error ~0: mode=add, td1coef, s3x5.

Uso: python scripts/calibrar_acero.py
"""
import os
import subprocess
import sys
import tempfile

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from etl.core import seasonal as S  # noqa: E402

XLSX = os.path.join(os.path.dirname(__file__), os.pardir,
                    "etl", "datasets", "acero", "data", "Acero.xlsx")


def load():
    ws = openpyxl.load_workbook(XLSX, data_only=True)["Hoja1"]
    dates, obs, ref = [], [], {}
    for r in range(2, ws.max_row + 1):
        d, v, de = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value
        if d is None or v is None:
            continue
        dd = d.date()
        dates.append(dd)
        obs.append(float(v))
        if de is not None:
            ref[(dd.year, dd.month)] = float(de)
    return dates, obs, ref


def run(x13, dates, obs, ref, mode, td, sma):
    wd = tempfile.mkdtemp(prefix="cal_acero_")
    S._write_spc(os.path.join(wd, "serie.spc"), dates, obs, mode=mode, td=td, seasonalma=sma)
    subprocess.run([x13, "serie"], cwd=wd, capture_output=True, text=True, timeout=120)
    p = os.path.join(wd, "serie.d11")
    if not os.path.isfile(p):
        return None
    d11 = {(d.year, d.month): v for d, v in S._parse_d11(p)}
    common = [k for k in ref if k in d11]
    if not common:
        return None
    pct = [abs(d11[k] - ref[k]) / abs(ref[k]) * 100 for k in common]
    diffs = [abs(d11[k] - ref[k]) for k in common]
    return dict(mode=mode, td=td, sma=sma, arima=S._arima_model(wd, "serie"),
                meanabs=sum(diffs) / len(diffs), maxabs=max(diffs), maxpct=max(pct))


def main():
    x13 = S._x13_binary()
    if not x13:
        sys.exit("x13as no encontrado (setear X13PATH)")
    dates, obs, ref = load()
    print(f"serie: {len(obs)} meses {dates[0]:%Y-%m}..{dates[-1]:%Y-%m} | ref: {len(ref)}")
    results = []
    for mode in ("add", "mult", "auto"):
        for td in ("none", "td1coef", "td"):
            for sma in ("s3x5", "s3x3", "s3x9"):
                try:
                    r = run(x13, dates, obs, ref, mode, td, sma)
                except Exception:
                    r = None
                if r:
                    results.append(r)
    results.sort(key=lambda r: r["meanabs"])
    print(f"\n{'mode':5} {'td':8} {'sma':5} {'arima':16} {'meanabs':>9} {'maxabs':>9} {'maxpct%':>9}")
    for r in results[:8]:
        print(f"{r['mode']:5} {r['td']:8} {r['sma']:5} {str(r['arima']):16} "
              f"{r['meanabs']:9.4f} {r['maxabs']:9.4f} {r['maxpct']:9.4f}")


if __name__ == "__main__":
    main()
