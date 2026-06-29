"""Harness de calibración del X-13 contra la referencia de Juan (produccion).

Objetivo: encontrar qué configuración de X-13 (modo, modelo, outliers, filtro estacional)
reproduce los valores desestacionalizados de Juan a partir de los originales.

Lee `autos_prod.xlsx` (3 columnas sin header: fecha | original | desest_juan), y para
cada variante de spec corre el binario x13as sobre los ORIGINALES de Juan, lee la serie
ajustada y la compara contra la columna desest_juan. Reporta:
  - err_medio   : error absoluto medio sobre TODOS los meses
  - err_exCOVID : error medio EXCLUYENDO 2020-2021 (si esto se va a ~0, matcheamos salvo COVID)
  - err_max @   : peor error y en qué mes cae
  - abr2020     : el valor en abril-2020 (Juan da -1914; es el "tell")

Iteración 2: ninguna de las variantes estándar matcheó (las aditivas dan abr2020 ~+690,
Juan da -1914), así que ahora barremos el FILTRO ESTACIONAL del X-11 (seasonalma), que es
la palanca que más mueve los factores estacionales.

Uso (en el server, venv activado, X13PATH seteado):
    python scripts/calibrar.py [ruta_al_xlsx]
NO toca la base. Deja la salida de cada variante en /tmp/calibrar/<n>/.
"""
from __future__ import annotations

import os
import re
import subprocess
import sys
from datetime import date

import openpyxl

OUT_BASE = "/tmp/calibrar"
APR2020 = (2020, 4)
COVID_YEARS = {2020, 2021}
_MODEL_RE = re.compile(r"\(\s*\d+\s+\d+\s+\d+\s*\)\s*\(\s*\d+\s+\d+\s+\d+\s*\)")
_TAG_RE = re.compile(r"<[^>]*>")


def find_binary() -> str:
    p = os.environ.get("X13PATH")
    cands = []
    if p:
        if os.path.isfile(p):
            return p
        cands += [os.path.join(p, n) for n in ("x13as_html", "x13as", "x13ashtml")]
    cands.append(os.path.expanduser("~/x13/x13as/x13as_html"))
    for c in cands:
        if os.path.isfile(c):
            return c
    sys.exit(f"No encontré el binario x13as. Seteá X13PATH. Probé: {cands}")


def read_juan(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    dates, orig, des = [], [], []
    for d, o, s in ws.iter_rows(values_only=True):
        if d is None or o is None:
            continue
        dd = d.date() if hasattr(d, "date") else d
        dates.append(date(dd.year, dd.month, 1))
        orig.append(float(o))
        des.append(float(s) if s is not None else None)
    wb.close()
    return dates, orig, des


def interp_nonpositive(values):
    v = list(values)
    n = len(v)
    for i in range(n):
        if v[i] is None or v[i] <= 0:
            j = i - 1
            while j >= 0 and (v[j] is None or v[j] <= 0):
                j -= 1
            k = i + 1
            while k < n and (values[k] is None or values[k] <= 0):
                k += 1
            left = v[j] if j >= 0 else None
            right = values[k] if k < n else None
            if left is not None and right is not None:
                v[i] = left + (right - left) * (i - j) / (k - j)
            else:
                v[i] = left if left is not None else (right if right is not None else 1.0)
    return v


def build_spc(dates, values, opt) -> str:
    y, m = dates[0].year, dates[0].month
    nums = [f"{x:.4f}" for x in values]
    blocks = ["  " + " ".join(nums[i:i + 10]) for i in range(0, len(nums), 10)]
    data = "\n".join(blocks)
    lines = [f'series{{ title="serie" start={y}.{m:02d} period=12\n data=(\n{data}\n ) }}']
    if opt.get("transform"):
        lines.append(f'transform{{ function={opt["transform"]} }}')
    if opt.get("reg"):
        lines.append(f"regression{{ variables=({opt['reg']}) }}")
    model = opt.get("model")
    if model == "automdl":
        lines.append("automdl{ }")
    elif model:
        lines.append(f"arima{{ model={model} }}")
    if opt.get("outlier"):
        crit = opt.get("critical")
        lines.append("outlier{ " + (f"critical={crit} " if crit else "") + "}")
    if opt.get("decomp", "x11") == "x11":
        parts = []
        if opt.get("mode"):
            parts.append(f"mode={opt['mode']}")
        if opt.get("seasonalma"):
            parts.append(f"seasonalma={opt['seasonalma']}")
        if opt.get("trendma"):
            parts.append(f"trendma={opt['trendma']}")
        parts.append("save=(d11)")
        lines.append("x11{ " + " ".join(parts) + " }")
        if opt.get("x11reg"):  # trading-day estimado por el X-11 (forma "clásica")
            lines.append(f"x11regression{{ variables=({opt['x11reg']}) }}")
    else:
        lines.append("seats{ save=(s11) }")
    return "\n".join(lines) + "\n"


def parse_table(path):
    out = {}
    with open(path) as f:
        for ln in f:
            parts = ln.split()
            if len(parts) < 2:
                continue
            ym = parts[0]
            if len(ym) == 6 and ym.isdigit():
                try:
                    out[(int(ym[:4]), int(ym[4:6]))] = float(parts[1])
                except ValueError:
                    pass
    return out


def arima_from_html(workdir):
    p = os.path.join(workdir, "serie.html")
    if not os.path.isfile(p):
        return None
    with open(p, encoding="utf-8", errors="ignore") as f:
        text = _TAG_RE.sub("", f.read())
    for label in ("Final automatic model choice", "ARIMA Model:"):
        i = text.find(label)
        if i != -1:
            mm = _MODEL_RE.search(text, i)
            if mm:
                return re.sub(r"\s+", " ", mm.group(0))
    return None


def run_variant(binary, dates, values, opt, workdir):
    os.makedirs(workdir, exist_ok=True)
    with open(os.path.join(workdir, "serie.spc"), "w") as f:
        f.write(build_spc(dates, values, opt))
    subprocess.run([binary, "serie"], cwd=workdir, capture_output=True, text=True, timeout=180)
    ext = ".d11" if opt.get("decomp", "x11") == "x11" else ".s11"
    p = os.path.join(workdir, "serie" + ext)
    return parse_table(p) if os.path.isfile(p) else None


def compare(series, dates, desest):
    errs, errs_ex = [], []
    apr, mx = None, (-1.0, None)
    for d, s in zip(dates, desest):
        if s is None:
            continue
        got = series.get((d.year, d.month))
        if got is None:
            continue
        e = abs(got - s)
        errs.append(e)
        if d.year not in COVID_YEARS:
            errs_ex.append(e)
        if e > mx[0]:
            mx = (e, f"{d.year}-{d.month:02d}")
        if (d.year, d.month) == APR2020:
            apr = got
    if not errs:
        return None
    return dict(mean=sum(errs) / len(errs),
                mean_ex=(sum(errs_ex) / len(errs_ex)) if errs_ex else float("nan"),
                mx=mx[0], mx_at=mx[1], apr=apr)


# Iteración 4: el TRADING-DAY bajó el error a la mitad (~1144 -> 484). Afinamos: TD vía
# regARIMA vs vía X-11 (x11regression, la forma clásica), variantes de TD, sensibilidad de
# outliers (critical) y modelos fijos. base = aditivo + automdl + td + outlier.
GRID = [
    dict(label="ref: TD(regARIMA) + outlier  [mejor iter3]", transform="none", model="automdl", outlier=True, mode="add", reg="td"),
    dict(label="TD via X-11 (x11regression) + outlier", transform="none", model="automdl", outlier=True, mode="add", x11reg="td"),
    dict(label="TD via X-11 (x11regression), SIN outlier", transform="none", model="automdl", mode="add", x11reg="td"),
    dict(label="TD1coef(regARIMA) + outlier", transform="none", model="automdl", outlier=True, mode="add", reg="td1coef"),
    dict(label="TD(regARIMA) + outlier critical=3", transform="none", model="automdl", outlier=True, critical=3, mode="add", reg="td"),
    dict(label="TD(regARIMA) + outlier critical=5", transform="none", model="automdl", outlier=True, critical=5, mode="add", reg="td"),
    dict(label="TD(regARIMA) + outlier + seasonalma=s3x5", transform="none", model="automdl", outlier=True, mode="add", reg="td", seasonalma="s3x5"),
    dict(label="TD(regARIMA) + ARIMA(0 1 1)(0 1 1) + outlier", transform="none", model="(0 1 1)(0 1 1)", outlier=True, mode="add", reg="td"),
]


def main(argv=None):
    argv = sys.argv[1:] if argv is None else argv
    xlsx = argv[0] if argv else os.path.join(os.path.dirname(__file__), "..", "autos_prod.xlsx")
    xlsx = os.path.abspath(xlsx)
    if not os.path.isfile(xlsx):
        sys.exit(f"No encontré el xlsx: {xlsx}")
    binary = find_binary()
    dates, orig, desest = read_juan(xlsx)
    juan_apr = next((s for d, s in zip(dates, desest) if (d.year, d.month) == APR2020), None)
    print(f"binario: {binary}")
    print(f"datos: {len(dates)} meses {dates[0]}..{dates[-1]} | desest_juan abr-2020 = {juan_apr:.0f}")
    print(f"(err_exCOVID excluye 2020-2021; si tiende a 0, matcheamos salvo COVID)\n")

    results = []
    for i, opt in enumerate(GRID):
        workdir = os.path.join(OUT_BASE, str(i))
        values = interp_nonpositive(orig) if opt.get("positive") else orig
        try:
            series = run_variant(binary, dates, values, opt, workdir)
        except Exception as e:
            print(f"[{i}] {opt['label']:46} ERROR: {e}")
            continue
        if not series:
            print(f"[{i}] {opt['label']:46} sin salida (ver {workdir}/serie_err.html)")
            continue
        c = compare(series, dates, desest)
        arima = arima_from_html(workdir) if opt.get("model") == "automdl" else (opt.get("model") or "x11")
        results.append((c, opt["label"], arima, i))

    print("\n==== RESULTADOS (orden por err_exCOVID; menor = mejor match a Juan) ====")
    print(f"{'err_medio':>10} {'err_exCOVID':>12} {'err_max':>10} {'@mes':>8} {'abr2020':>9}  modelo          variante")
    for c, label, arima, i in sorted(results, key=lambda r: (r[0]['mean_ex'])):
        print(f"{c['mean']:10.1f} {c['mean_ex']:12.1f} {c['mx']:10.0f} {c['mx_at']:>8} "
              f"{c['apr']:9.0f}  {str(arima):14}  [{i}] {label}")

    if results:
        best = min(results, key=lambda r: r[0]['mean_ex'])
        c = best[0]
        print(f"\nMEJOR (fuera de COVID): [{best[3]}] {best[1]}")
        print(f"  err_exCOVID={c['mean_ex']:.1f}  err_medio={c['mean']:.1f}  abr2020={c['apr']:.0f} vs juan={juan_apr:.0f}")
        print(f"  spc/reporte: {OUT_BASE}/{best[3]}/serie.spc  y  serie.html")


if __name__ == "__main__":
    main()
