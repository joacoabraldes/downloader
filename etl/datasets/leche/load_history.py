"""Carga histórica (one-off) de producción de leche desde el Excel de referencia (leche.xlsx).

El Excel tiene una hoja "Mensual" con columnas `date | valor | desest`: `valor` es la
producción mensual observada (litros, desde 2015-01) y `desest` es la referencia de X-13 (no
se ingesta; sirve para calibrar). Se carga `valor` con estado=NULL (histórico). Idempotente:
usa insert_if_changed. El PPV mensual (run.py) re-publica la serie como 'definitivo', que en
la vista _actual tiene prioridad sobre este histórico.
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

import openpyxl

from etl.core import db, report
from . import config

DEFAULT_XLSX = Path(__file__).parent / "data" / "leche.xlsx"
SHEET = "Mensual"
FUENTE = "excel historico"


def read_rows(path: str) -> list[tuple[dt.date, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    rows: list[tuple[dt.date, float]] = []
    for r in range(1, ws.max_row + 1):  # col A = date, col B = valor (litros), col C = desest
        d = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if isinstance(d, dt.datetime) and isinstance(v, (int, float)):
            rows.append((d.date(), float(v)))
    wb.close()
    return rows


def main(argv=None) -> None:
    ap = argparse.ArgumentParser(prog="etl leche load-history",
                                 description="Carga histórica del Excel de leche (one-off).")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="ruta al Excel")
    ap.add_argument("--force", action="store_true", help="re-insertar aunque no cambie")
    args = ap.parse_args(argv)

    rows = read_rows(args.xlsx)
    if not rows:
        print("No se leyeron filas del Excel.", file=sys.stderr)
        sys.exit(1)

    rep = report.Report("leche", "load-history")
    rep.info(f"Excel: {len(rows)} filas | rango: {rows[0][0]:%Y-%m}..{rows[-1][0]:%Y-%m}")
    conn = db.get_conn()
    try:
        for date, valor in rows:
            rep.tally(db.insert_if_changed(
                conn, table=config.TABLE, key_cols=config.KEY_COLS,
                key_vals=[config.MAIN_SERIE, date], value_cols=config.VALUE_COLS,
                row={"valor": valor}, estado=None, fuente=FUENTE, force=args.force,
            ))
    finally:
        conn.close()
    rep.summary()


if __name__ == "__main__":
    main()
