"""Fuente MAGyP (lechería): baja el Excel de producción mensual y parsea los litros.

La página de estadística primaria publica el Excel `PPV021_PPV022.xlsx` con **URL fija**
(MAGyP pisa el mismo archivo cada mes), así que el incremental baja esa URL directamente. Si
algún día cambiara, hay un fallback que scrapea la página buscando el link a ese archivo.

El Excel tiene una hoja "Mensual" con un encabezado y luego filas `MES | Litros` (la fecha en
la 2ª columna, los litros en la 3ª), producción nacional desde 2015.
"""
from __future__ import annotations

import datetime as dt
import io

import openpyxl
import requests
from bs4 import BeautifulSoup

BASE = "https://www.magyp.gob.ar/sitio/areas/ss_lecheria/estadisticas/_01_primaria"
XLSX_URL = f"{BASE}/_archivos/PPV021_PPV022.xlsx"
PAGE = f"{BASE}/index.php"
HEADERS = {"User-Agent": "Mozilla/5.0 (leche ETL)"}
TIMEOUT = 90


def _resolve_url() -> str:
    """URL del PPV: la fija; si 404, scrapea la página buscando el link a ese archivo."""
    r = requests.head(XLSX_URL, headers=HEADERS, timeout=30, verify=False,
                      allow_redirects=True)
    if r.status_code == 200:
        return XLSX_URL
    page = requests.get(PAGE, headers=HEADERS, timeout=TIMEOUT, verify=False)
    page.raise_for_status()
    soup = BeautifulSoup(page.content, "lxml")
    for a in soup.find_all("a", href=True):
        if "PPV021_PPV022" in a["href"] and a["href"].lower().endswith(".xlsx"):
            return requests.compat.urljoin(PAGE, a["href"])
    return XLSX_URL  # último recurso: probar igual la fija


def download(url: str) -> bytes:
    """Baja el archivo (verify=False: el cert de magyp.gob.ar no valida en el server)."""
    r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, verify=False)
    r.raise_for_status()
    return r.content


def parse_produccion(xlsx_bytes: bytes) -> dict[dt.date, float]:
    """{date(primer día del mes): litros} desde la hoja 'Mensual' del Excel PPV."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Mensual"] if "Mensual" in wb.sheetnames else wb[wb.sheetnames[0]]
    out: dict[dt.date, float] = {}
    for r in range(1, ws.max_row + 1):
        for col in (1, 2):  # la fecha puede estar en la 1ª o 2ª columna según el layout
            d = ws.cell(r, col).value
            v = ws.cell(r, col + 1).value
            if isinstance(d, dt.datetime) and isinstance(v, (int, float)):
                out[dt.date(d.year, d.month, 1)] = float(v)
                break
    wb.close()
    return out


def get_latest() -> tuple[dict[dt.date, float], str] | None:
    """({date: litros}, url) del Excel PPV, o None si no se parseó ninguna fila."""
    url = _resolve_url()
    data = parse_produccion(download(url))
    return (data, url) if data else None


if __name__ == "__main__":  # smoke test
    import urllib3
    urllib3.disable_warnings()
    res = get_latest()
    if res:
        data, url = res
        print(url)
        for d in sorted(data)[-6:]:
            print(f"  {d:%Y-%m}  {data[d]:,.0f}")
