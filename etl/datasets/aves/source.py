"""Fuente MAGyP: descubre el último Excel de indicadores avícolas y parsea la faena mensual.

La página de "Carne Aviar" no tiene tablas: los datos están en archivos. El más útil es
`Indicadores de Oferta y Demanda 2016-<año>.xlsx`, que trae la faena mensual (col "Faena
SENASA", miles de cabezas) de todos los años en un solo archivo y se actualiza cada mes.

El xlsx está apilado por año: una fila con el año (col A = '2016') y debajo 12 filas de mes
(col A = 'Enero'...). De cada fila de mes tomamos la col B = faena. El link del xlsx cambia
de carpeta/nombre por año, así que se scrapea la página y se elige el de mayor año.

**Fallback en PDF**: el xlsx de indicadores se actualiza más tarde que el PDF `Faena Avícola
<año>.pdf` de la misma página (jul-2026: el PDF ya tenía junio y el xlsx llegaba hasta mayo).
El PDF trae la misma serie pero **redondeada a la unidad** (67.120 vs 67119.556 del xlsx), así
que sirve para adelantar el mes como `provisorio`, nunca para pisar al xlsx. Ver
`find_latest_faena_pdf` / `parse_faena_pdf` y el uso en `run.py`.
"""
from __future__ import annotations

import datetime as dt
import io
import re
import unicodedata
import urllib.parse

import openpyxl
import pdfplumber
from bs4 import BeautifulSoup

from etl.core import http
from . import config

PAGE = "https://www.magyp.gob.ar/sitio/areas/aves/estadistica/carne/index.php"
HEADERS = {"User-Agent": "Mozilla/5.0 (aves ETL)"}
TIMEOUT = 90

_MESES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
          "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
          "noviembre": 11, "diciembre": 12}
FAENA_COL = 2  # col B del xlsx = "Faena SENASA" (miles de cabezas)

# Un año de 4 dígitos 20xx que no sea parte de un número más largo. Los nombres de MAGyP
# arrancan con un prefijo tipo `000000_` / `260000_`, así que un `\d{4}` pelado devuelve años
# fantasma ('0000', '2600'); esto sólo matchea 20xx aislado.
_ANIO = re.compile(r"(?<!\d)(20\d{2})(?!\d)")


def _sin_acentos(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower()


def _anio_de_href(href: str) -> int:
    """Mayor año 20xx del href COMPLETO (incluye la carpeta), o 0.

    Se mira el href entero y no sólo el nombre porque el archivo de 2025 se llama
    `000000_Faena Avícola.pdf`, sin año: el único lugar donde figura es la carpeta
    (`.../250000_2025/`).
    """
    years = [int(y) for y in _ANIO.findall(href)]
    return max(years) if years else 0


def find_latest_indicadores_xlsx() -> str | None:
    """URL absoluta del 'Indicadores de Oferta y Demanda' .xlsx de mayor año, o None."""
    r = http.fetch(PAGE, headers=HEADERS, timeout=TIMEOUT, verify=False)
    soup = BeautifulSoup(r.content, "lxml")
    best = None
    for a in soup.find_all("a", href=True):
        href = a["href"]
        name = href.rsplit("/", 1)[-1].lower()
        if "indicadores de oferta y demanda" not in name or not name.endswith(".xlsx"):
            continue
        years = [int(y) for y in re.findall(r"\d{4}", name)]
        key = max(years) if years else 0
        if best is None or key > best[0]:
            best = (key, urllib.parse.urljoin(PAGE, urllib.parse.quote(href)))
    return best[1] if best else None


def find_latest_faena_pdf() -> str | None:
    """URL absoluta del PDF de faena avícola nacional de mayor año, o None.

    El nombre cambia todos los años ('Faena Avícola 2026.pdf', 'Faena Avícola.pdf',
    'Faena de aves 2024.pdf'), así que se pide 'faena' + 'avicola'/'aves'. El
    **'provincial' se excluye explícitamente**: `Faena Provincial <años>.pdf` vive en la misma
    carpeta, matchea 'faena' y es otra serie (faena abierta por provincia).
    """
    r = http.fetch(PAGE, headers=HEADERS, timeout=TIMEOUT, verify=False)
    soup = BeautifulSoup(r.content, "lxml")
    best = None
    for a in soup.find_all("a", href=True):
        href = a["href"]
        name = _sin_acentos(href.rsplit("/", 1)[-1])
        if not name.endswith(".pdf") or "faena" not in name or "provincial" in name:
            continue
        if "avicola" not in name and "aves" not in name:
            continue
        key = _anio_de_href(_sin_acentos(href))
        if best is None or key > best[0]:
            best = (key, urllib.parse.urljoin(PAGE, urllib.parse.quote(href)))
    return best[1] if best else None


def download(url: str) -> bytes:
    """Baja el archivo (verify=False: el cert de magyp.gob.ar no valida en el server)."""
    r = http.fetch(url, headers=HEADERS, timeout=TIMEOUT, verify=False)
    return r.content


def parse_faena(xlsx_bytes: bytes) -> dict[dt.date, float]:
    """{date(primer día del mes): faena} desde el xlsx apilado por año."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb[wb.sheetnames[0]]
    out: dict[dt.date, float] = {}
    year = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        s = str(a).strip()
        if s.isdigit() and len(s) == 4:  # fila de año -> abre el bloque
            year = int(s)
            continue
        mes = _MESES.get(s.lower())
        b = ws.cell(r, FAENA_COL).value
        if mes and year and b is not None:
            out[dt.date(year, mes, 1)] = float(b)
    wb.close()
    return out


def get_latest() -> tuple[dict[dt.date, float], str] | None:
    """({date: faena}, url) del último xlsx de indicadores, o None si no se encontró."""
    url = find_latest_indicadores_xlsx()
    if not url:
        return None
    return parse_faena(download(url)), url


# --- Fallback: PDF de faena --------------------------------------------------------------
# Layout del PDF (1 página, dos años lado a lado):
#
#   FAENA AVICOLA
#   Año 2026/2025
#   2025 2026
#   Mes            Variación %
#   Enero    67.120  60.962  -9,2      <- los dos años + variación
#   ...
#   Julio    66.990                    <- SOLO el año viejo (el nuevo no se publicó)
#   Total a Junio  366.199 358.979 -2,0
#   Total anual    752.562
#
# Las filas de un solo número son la trampa: ese número es del año VIEJO, no del nuevo. Un
# parser que tome "el último número de la fila" carga julio-2025 como julio-2026 e inventa un
# dato. Por eso el orden de columnas se lee del encabezado y la cantidad de valores decide a
# qué año va cada uno; cualquier fila que no encaje se descarta.

# Valor con separador de miles por punto y sin decimales ('67.120'). La variación siempre
# lleva coma decimal ('-9,2'), así que este patrón la deja afuera sola.
_VALOR = re.compile(r"^\d{1,3}(?:\.\d{3})*$")
# Encabezado de columnas: una línea con exactamente dos años, en el orden en que aparecen.
_COLS = re.compile(r"^\s*(20\d{2})\s+(20\d{2})\s*$", re.M)


def _a_float(tok: str) -> float:
    return float(tok.replace(".", ""))


def parse_faena_pdf(pdf_bytes: bytes) -> dict[dt.date, float]:
    """{date(primer día del mes): faena} desde el PDF de faena avícola.

    Los valores vienen redondeados a la unidad (miles de cabezas), a diferencia del xlsx que
    trae 3 decimales. Levanta ValueError si no se puede determinar el orden de las columnas:
    antes que adivinar el año, se corta.
    """
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        texto = "\n".join(p.extract_text() or "" for p in pdf.pages)

    cols = _COLS.search(texto)
    if not cols:
        raise ValueError("no se encontró el encabezado de columnas (dos años) en el PDF")
    anios = (int(cols.group(1)), int(cols.group(2)))  # orden izquierda -> derecha

    out: dict[dt.date, float] = {}
    for linea in texto.splitlines():
        toks = linea.split()
        if not toks:
            continue
        mes = _MESES.get(_sin_acentos(toks[0]))
        if not mes:  # descarta 'Total a Junio' / 'Total anual': no arrancan con el mes
            continue
        valores = [t for t in toks[1:] if _VALOR.match(t)]
        if len(valores) == 2:      # los dos años
            for anio, tok in zip(anios, valores):
                out[dt.date(anio, mes, 1)] = _a_float(tok)
        elif len(valores) == 1:    # sólo la primera columna (año viejo)
            out[dt.date(anios[0], mes, 1)] = _a_float(valores[0])
        # cualquier otra cantidad: fila inesperada, se ignora (no se adivina)
    return out


def get_latest_pdf() -> tuple[dict[dt.date, float], str] | None:
    """({date: faena}, url) del PDF de faena, o None si no se encontró el link."""
    url = find_latest_faena_pdf()
    if not url:
        return None
    return parse_faena_pdf(download(url)), url


if __name__ == "__main__":  # smoke test
    import urllib3
    urllib3.disable_warnings()
    res = get_latest()
    if res:
        data, url = res
        print(url)
        for d in sorted(data)[-6:]:
            print(f"  {d:%Y-%m}  {data[d]}")
