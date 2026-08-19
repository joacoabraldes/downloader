"""Fuente CAMMESA: baja el Excel "Demanda Mensual" y parsea el cuadro de demanda por tipo.

El link "Demanda Mensual" de la página de síntesis del MEM devuelve el .xlsx directo (no un
zip), re-publicado entero cada mes. Hoja única 'DEMANDA', unidad MWh, tabla horizontal: la
fila header (col A = 'TIPO DEMANDA') tiene las fechas mensuales en las columnas B+, y debajo
una fila por tipo de demanda (col A = etiqueta). Se parsea POR ETIQUETA (ver config), se
convierte a GWh (÷1000) y se arma además la serie derivada `no_residencial`.
"""
from __future__ import annotations

import datetime as dt
import io

import openpyxl

from etl.core import http
from . import config

# Link "Demanda Mensual" (WP Download Manager, wpdmdl estable) de estadistica-informe-sintesis-mem.
URL = "https://cammesaweb.cammesa.com/download/demanda-mensual/?wpdmdl=41426"
REFERER = "https://cammesaweb.cammesa.com/estadistica-informe-sintesis-mem/"
HEADERS = {"User-Agent": "Mozilla/5.0 (demanda_energia ETL)", "Referer": REFERER}
TIMEOUT = 180
SHEET = "DEMANDA"
HEADER_LABEL = "tipo demanda"  # col A de la fila de encabezado (fechas)
MAX_HEADER_SCAN = 40           # las filas del cuadro están arriba de todo


def download(url: str = URL) -> bytes:
    r = http.fetch(url, headers=HEADERS, timeout=TIMEOUT)
    return r.content


def parse(xlsx_bytes: bytes) -> dict[str, dict[dt.date, float]]:
    """{serie: {date(primer día del mes): valor_gwh}} desde la hoja DEMANDA.

    Incluye la serie derivada `no_residencial` (suma de sus componentes, ver config).
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN, values_only=True))
    wb.close()

    # Fila de encabezado: col A == 'TIPO DEMANDA'. Sus celdas datetime son los meses.
    header = next((row for row in rows
                   if row and isinstance(row[0], str)
                   and row[0].strip().lower() == HEADER_LABEL), None)
    if header is None:
        return {}
    monthcol = {i: dt.date(v.year, v.month, 1)
                for i, v in enumerate(header) if isinstance(v, dt.datetime)}
    if not monthcol:
        return {}

    out: dict[str, dict[dt.date, float]] = {}
    for row in rows:
        label = row[0] if row else None
        if not isinstance(label, str):
            continue
        serie = config.LABEL_TO_SERIE.get(label.strip().lower())
        if serie is None:
            continue
        serie_data = out.setdefault(serie, {})
        for i, date in monthcol.items():
            v = row[i] if i < len(row) else None
            if isinstance(v, (int, float)):
                serie_data[date] = float(v) / 1000.0  # MWh -> GWh

    # Serie derivada: no residencial total = suma de sus componentes, mes a mes.
    parts = [out.get(p, {}) for p in config.NO_RESIDENCIAL_PARTS]
    dates = set().union(*parts) if parts else set()
    no_res = {}
    for d in dates:
        if all(d in p for p in parts):  # solo meses con todos los componentes
            no_res[d] = sum(p[d] for p in parts)
    if no_res:
        out[config.MAIN_SERIE] = no_res
    return out


def get_latest() -> tuple[dict[str, dict[dt.date, float]], str] | None:
    """({serie: {date: valor_gwh}}, url) del Excel de CAMMESA, o None si no se parseó nada."""
    data = parse(download())
    if not data:
        return None
    return data, URL


if __name__ == "__main__":  # smoke test
    res = get_latest()
    if res:
        data, url = res
        print(url)
        for serie in sorted(data):
            d = data[serie]
            print(f"  {serie:22} {len(d):3} meses {min(d):%Y-%m}..{max(d):%Y-%m}  "
                  f"ultimo={d[max(d)]:.2f}")
