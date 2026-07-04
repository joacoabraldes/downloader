"""Fuente MAGyP (bovinos): descubre el xls de faena/producción y parsea la producción mensual.

La página de información sectorial no linkea el xls de datos directamente: linkea un PDF
"Tablero de Faena Bovina" que, **adentro**, tiene el hipervínculo al xls mensual
(`Faena_Bovina_<años>_mensual..xls`). El nombre del xls cambia con el rango de años, pero el
link embebido en el PDF siempre apunta al vigente. Así que la cadena es:
página → Tablero PDF → hipervínculo embebido → xls.

El archivo tiene una fila de encabezado con "Mes/Año" y "Producción (en miles de toneladas res
con hueso)". Se ubican esas columnas por texto (no por índice fijo) y se parsea la serie
mensual de producción desde 2019. Hoy MAGyP lo publica en `.xls` viejo (se lee con `xlrd`),
pero el parser detecta el formato por los magic bytes y cae a `openpyxl` si algún día pasan a
`.xlsx` moderno.
"""
from __future__ import annotations

import datetime as dt
import io

import openpyxl
import pdfplumber
import requests
import xlrd
from bs4 import BeautifulSoup

PAGE = "https://www.magyp.gob.ar/sitio/areas/bovinos/informacion_sectorial/"
HEADERS = {"User-Agent": "Mozilla/5.0 (bovinos ETL)"}
TIMEOUT = 90


def _find_tablero_pdf() -> str:
    """URL del PDF 'Tablero de Faena Bovina' (linkeado en la página de info sectorial)."""
    r = requests.get(PAGE, headers=HEADERS, timeout=TIMEOUT, verify=False)
    r.raise_for_status()
    soup = BeautifulSoup(r.content, "lxml")
    for a in soup.find_all("a", href=True):
        name = a["href"].rsplit("/", 1)[-1].lower()
        if "faena_bovino" in name and name.endswith(".pdf"):
            return requests.compat.urljoin(PAGE, a["href"])
    raise RuntimeError("no se encontró el link al Tablero de Faena Bovina en la página")


def _find_xls_url(pdf_bytes: bytes) -> str:
    """URL del xls de faena, extraída del hipervínculo embebido en el Tablero PDF."""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for h in (page.hyperlinks or []):
                uri = h.get("uri", "")
                if "faena_bovina" in uri.lower() and ".xls" in uri.lower():
                    return uri
    raise RuntimeError("el Tablero PDF no tiene el hipervínculo al xls de faena")


def download(url: str) -> bytes:
    """Baja el archivo (verify=False: el cert de magyp.gob.ar no valida en el server)."""
    r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, verify=False)
    r.raise_for_status()
    return r.content


def _find_cols(cell, nrows: int, ncols: int) -> tuple[int, int, int]:
    """(fila de encabezado, col de fecha 'Mes/Año', col de producción) buscando por texto.

    `cell(r, c)` devuelve el valor de la celda (str/num/None) sea cual sea la lib de lectura.
    """
    for r in range(min(10, nrows)):
        fecha_col = prod_col = None
        for c in range(ncols):
            txt = str(cell(r, c) or "").lower()
            if "mes" in txt and ("año" in txt or "ano" in txt):
                fecha_col = c
            if "producci" in txt and "res con hueso" in txt:
                prod_col = c
        if fecha_col is not None and prod_col is not None:
            return r, fecha_col, prod_col
    raise RuntimeError("no se ubicaron las columnas 'Mes/Año' y 'Producción' en el archivo")


def _parse_xls(blob: bytes) -> dict[dt.date, float]:
    """Parseo del formato .xls viejo (OLE2) con xlrd. La fecha viene como serial de Excel."""
    wb = xlrd.open_workbook(file_contents=blob)
    sh = wb.sheet_by_index(0)
    header, fecha_col, prod_col = _find_cols(sh.cell_value, sh.nrows, sh.ncols)
    out: dict[dt.date, float] = {}
    for r in range(header + 1, sh.nrows):
        f, v = sh.cell_value(r, fecha_col), sh.cell_value(r, prod_col)
        if not isinstance(f, float) or f <= 1000 or not isinstance(v, (int, float)) or not v:
            continue
        y, m, _ = xlrd.xldate_as_tuple(f, wb.datemode)[:3]
        out[dt.date(y, m, 1)] = float(v)
    return out


def _parse_xlsx(blob: bytes) -> dict[dt.date, float]:
    """Parseo del formato .xlsx moderno con openpyxl (fallback). La fecha es un datetime."""
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    ws = wb[wb.sheetnames[0]]
    cell = lambda r, c: ws.cell(r + 1, c + 1).value  # openpyxl es 1-based  # noqa: E731
    header, fecha_col, prod_col = _find_cols(cell, ws.max_row, ws.max_column)
    out: dict[dt.date, float] = {}
    for r in range(header + 1, ws.max_row):
        f, v = cell(r, fecha_col), cell(r, prod_col)
        if not isinstance(f, dt.datetime) or not isinstance(v, (int, float)) or not v:
            continue
        out[dt.date(f.year, f.month, 1)] = float(v)
    wb.close()
    return out


def parse_produccion(blob: bytes) -> dict[dt.date, float]:
    """{date(primer día del mes): produccion (miles tn res con hueso)}.

    Detecta el formato por los magic bytes: ZIP (`PK`) = .xlsx moderno (openpyxl); si no,
    OLE2 = .xls viejo (xlrd), que es lo que MAGyP publica hoy.
    """
    return _parse_xlsx(blob) if blob[:2] == b"PK" else _parse_xls(blob)


def get_latest() -> tuple[dict[dt.date, float], str] | None:
    """({date: produccion}, url_xls) siguiendo la cadena página → PDF → xls, o None."""
    xls_url = _find_xls_url(download(_find_tablero_pdf()))
    data = parse_produccion(download(xls_url))
    return (data, xls_url) if data else None


if __name__ == "__main__":  # smoke test
    import urllib3
    urllib3.disable_warnings()
    res = get_latest()
    if res:
        data, url = res
        print(url)
        for d in sorted(data)[-6:]:
            print(f"  {d:%Y-%m}  {data[d]:.3f}")
