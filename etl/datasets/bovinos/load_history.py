"""Carga histórica (one-off) de producción bovina. Dos orígenes, los dos hacia atrás.

**Excel de referencia** (`Bovinos.xlsx`, default): una hoja con columnas `date | valor | desest`.
`valor` es la producción mensual observada (miles tn res con hueso, desde 1998-01) y `desest` es
la referencia de X-13 (no se ingesta; sirve para calibrar). Se carga `valor` con estado=NULL
(histórico). El xls mensual de MAGyP (run.py) re-publica la serie 2019→ como 'definitivo', que en
la vista _actual tiene prioridad sobre este histórico.

**Planilla de MAGyP desde 1990** (`--magyp-1990`): MAGyP publica una segunda planilla, colgada de
otro PDF de la misma página, que arranca en 1990-01 y cubre los 8 años que al Excel de referencia
le faltan (1990-1997). Ver el bloque de comentarios en `source.py`.

    python -m etl bovinos load-history --magyp-1990

Ese modo **sólo inserta meses ANTERIORES al mínimo que ya está en la base**. No es una precaución
de más: las dos planillas de MAGyP se contradicen en 4 de los 87 meses que comparten (peor caso
2019-06, 2,75% de diferencia), así que tocar el solapamiento dejaría la serie flapeando entre dos
versiones —la corrida diaria la volvería a pisar al otro día—. Como efecto secundario la regla lo
vuelve idempotente por construcción: una segunda corrida no encuentra nada por debajo del mínimo
y termina sin escribir.

Los dos modos son one-off: el cron sólo corre `python -m etl bovinos`.
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

import openpyxl
import urllib3

from etl.core import db, report
from . import config, source

DEFAULT_XLSX = Path(__file__).parent / "data" / "Bovinos.xlsx"
SHEET = "Hoja1"
FUENTE = "excel historico"


def read_rows(path: str) -> list[tuple[dt.date, float]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    rows: list[tuple[dt.date, float]] = []
    for r in range(2, ws.max_row + 1):  # fila 1 = headers (date, valor, desest)
        d = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if isinstance(d, dt.datetime) and isinstance(v, (int, float)):
            rows.append((d.date(), float(v)))
    wb.close()
    return rows


def _cargar_magyp_1990(conn, rep, *, force: bool) -> None:
    """Inserta de la planilla de MAGyP sólo los meses ANTERIORES al mínimo que ya hay en la base.

    El corte por `min(date)` es la única salvaguarda que importa acá: sin él, esta planilla
    pisaría meses que la corrida diaria vuelve a escribir con los valores del otro archivo de
    MAGyP —que no coinciden— y la serie quedaría alternando entre dos versiones.
    """
    with conn.cursor() as cur:
        # Se excluye la desestacionalizada: es una serie derivada y arranca donde arranca la
        # observada, así que incluirla no cambiaría el mínimo, pero dejarlo explícito evita que
        # un cambio futuro en la X-13 mueva este corte sin que nadie lo note.
        cur.execute(f"select min(date) from {config.TABLE} "
                    "where estado is distinct from 'desestacionalizado'")
        minimo = cur.fetchone()[0]

    data, url = source.get_historico_magyp()
    rep.info(f"planilla MAGyP: {len(data)} meses | {min(data):%Y-%m}..{max(data):%Y-%m}")
    if minimo is None:
        rep.info("la tabla está vacía: entra la planilla completa")
        nuevos = sorted(data)
    else:
        nuevos = sorted(f for f in data if f < minimo)
        rep.info(f"mínimo en base: {minimo:%Y-%m} -> se cargan {len(nuevos)} meses anteriores "
                 f"(el resto se ignora a propósito, lo cubre el ETL diario)")
    if not nuevos:
        rep.info("no hay meses por debajo del mínimo: nada que hacer")
        return
    rep.info(f"fuente: {url}")
    for fecha in nuevos:
        rep.tally(db.insert_if_changed(
            conn, table=config.TABLE, key_cols=config.KEY_COLS,
            key_vals=[config.MAIN_SERIE, fecha], value_cols=config.VALUE_COLS,
            row={"valor": float(data[fecha])}, estado=None, fuente=url, force=force,
        ))


def main(argv=None) -> None:
    ap = argparse.ArgumentParser(prog="etl bovinos load-history",
                                 description="Carga histórica de bovinos (one-off).")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="ruta al Excel")
    ap.add_argument("--force", action="store_true", help="re-insertar aunque no cambie")
    ap.add_argument("--magyp-1990", action="store_true",
                    help="cargar el tramo viejo (1990-1997) desde la planilla de MAGyP en vez "
                         "del Excel de referencia; sólo inserta por debajo del mínimo actual")
    args = ap.parse_args(argv)

    rep = report.Report("bovinos", "load-history")
    conn = db.get_conn()
    try:
        if args.magyp_1990:
            urllib3.disable_warnings()  # cert de magyp.gob.ar (verify=False)
            _cargar_magyp_1990(conn, rep, force=args.force)
        else:
            rows = read_rows(args.xlsx)
            if not rows:
                print("No se leyeron filas del Excel.", file=sys.stderr)
                sys.exit(1)
            rep.info(f"Excel: {len(rows)} filas | rango: {rows[0][0]:%Y-%m}..{rows[-1][0]:%Y-%m}")
            for date, valor in rows:
                rep.tally(db.insert_if_changed(
                    conn, table=config.TABLE, key_cols=config.KEY_COLS,
                    key_vals=[config.MAIN_SERIE, date], value_cols=config.VALUE_COLS,
                    row={"valor": valor}, estado=None, fuente=FUENTE, force=args.force,
                ))
    finally:
        conn.close()
    rep.summary()


if __name__ == "__main__":
    main()
