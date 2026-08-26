"""Carga histórica (one-off) de los totales de petróleo y gas desde los Excel de referencia.

Cada Excel (`petroleo.xlsx`, `Gas.xlsx`) trae en `Hoja1` tres columnas SIN encabezado:

    A = fecha (primer día del mes)   B = observado   C = desestacionalizado de referencia

Se cargan 367 filas: 1996-01 → 2026-07. Sólo entra la columna B, con estado=NULL (histórico).
La columna C **no se ingesta**: es la referencia contra la que se calibró la receta de X-13
(ver `scripts/calibrar_hidrocarburos.py`), no un dato de la fuente.

Las hojas tienen basura debajo del rango de datos (bloques pegados a mano, filas de fórmulas),
por eso se corta en `FILAS` y no en `ws.max_row`.

El histórico cubre 1996→2026, pero de 2009 en adelante el `run.py` re-publica el mismo mes
como 'provisorio' con el dato primario de Superset, que en la vista `_actual` tiene prioridad.
Las dos fuentes coinciden con 0,000% en 2009-2013, así que el empalme no produce escalón; de
2019 en adelante el histórico corre ~1% por debajo (ver el docstring de `schema.sql`).

El desagregado por tipo de recurso no tiene histórico: arranca en 2009 y lo carga `run.py`.

Idempotente: usa insert_if_changed, así que re-correrlo no duplica.
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

import openpyxl

from etl.core import db, report
from . import config

DATA = Path(__file__).parent / "data"
FUENTE = "excel historico"

# Un Excel por serie. Los nombres son los que subió el usuario al repo (ojo con la mayúscula
# de Gas.xlsx: en Linux importa).
XLSX = {"petroleo": DATA / "petroleo.xlsx", "gas": DATA / "Gas.xlsx"}

HOJA = "Hoja1"
FILAS = 367  # 1996-01 .. 2026-07; debajo de eso las hojas tienen bloques sueltos y fórmulas


def read_series(path: str) -> list[tuple[dt.date, float]]:
    """Lee (date, observado) de Hoja1!A1:B<FILAS>. Sin encabezado: la fila 1 ya es dato."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[HOJA]
    out: list[tuple[dt.date, float]] = []
    for fecha, valor in ws.iter_rows(min_row=1, max_row=FILAS, max_col=2, values_only=True):
        if fecha is None or valor is None:
            continue
        fecha = fecha.date() if isinstance(fecha, dt.datetime) else fecha
        try:
            out.append((fecha, float(valor)))
        except (TypeError, ValueError):
            continue  # celda no numérica suelta: se saltea
    wb.close()
    return out


def main(argv=None) -> None:
    ap = argparse.ArgumentParser(prog="etl hidrocarburos load-history",
                                 description="Carga histórica de petroleo/gas (one-off).")
    ap.add_argument("--serie", choices=config.TOTALES,
                    help="cargar sólo una de las dos series (default: las dos)")
    ap.add_argument("--force", action="store_true", help="re-insertar aunque no cambie")
    args = ap.parse_args(argv)

    series = [args.serie] if args.serie else list(config.TOTALES)
    rep = report.Report("hidrocarburos", "load-history")
    conn = db.get_conn()
    try:
        for serie in series:
            path = XLSX[serie]
            if not path.is_file():
                rep.error(f"{serie}: no está {path}")
                continue
            rows = read_series(str(path))
            if not rows:
                rep.error(f"{serie}: no se leyeron filas de {path.name}")
                continue
            rep.info(f"{serie}: {len(rows)} filas | "
                     f"rango {rows[0][0]:%Y-%m}..{rows[-1][0]:%Y-%m} | {path.name}")
            for fecha, valor in rows:
                rep.tally(db.insert_if_changed(
                    conn, table=config.TABLE, key_cols=config.KEY_COLS,
                    key_vals=[serie, fecha], value_cols=config.VALUE_COLS,
                    row={"valor": valor}, estado=None, fuente=FUENTE, force=args.force,
                ))
    finally:
        conn.close()
    rep.summary()
    if not rep.counts["leidos"]:
        sys.exit(1)


if __name__ == "__main__":
    main()
